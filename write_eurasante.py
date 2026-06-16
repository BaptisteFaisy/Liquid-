import openpyxl

wb = openpyxl.load_workbook('c:/Users/gaiam/Desktop/Liquid-/Liquid+ Project/Sélection des startups/processus de sélection des startups GAIA v2.xlsx')
ws = wb['startups (1)']

# Colonnes:
# C1=décision, C2=nom, C3=email, C4=email et/ou linkedin, C5=founder,
# C6=offre, C7=stade, C10=accompagnée?, C11=trouvée sur?

# Startups Eurasanté depuis le 01/01/2024
# Format: (nom, email, linkedin_ou_contact, founder, offre, stade, accompagnee, trouvee_sur)
startups = [
    # ======= BIO-INCUBATEUR EURASANTÉ =======
    (
        "Careflai",
        "",
        "https://lille.eurasante.com/startups/",
        "",
        "Santé / e-santé - Plateforme de télé-expertise médicale assistée par IA (biologie médicale)",
        "En incubation - Bio-Incubateur Eurasanté (2024)",
        "oui - Eurasanté (Bio-Incubateur)",
        "eurasante.com"
    ),
    (
        "Lifebloom",
        "",
        "https://fr.linkedin.com/in/damiensanderroche",
        "Damien Roche (Founder & CEO)",
        "Santé / MedTech - Dispositif médical pour permettre aux personnes en fauteuil de remarcher (Lifebloom One)",
        "En incubation - Bio-Incubateur Eurasanté (actif 2024)",
        "oui - Eurasanté (Bio-Incubateur)",
        "eurasante.com"
    ),
    (
        "iAVC",
        "",
        "https://www.linkedin.com/company/iavc-romain | https://iavc.fr",
        "Pierre-Olivier Bussière, Adrien Rousseaux",
        "Santé / e-santé - Application ROMAIN de géolocalisation des patients AVC en temps réel",
        "En incubation - Bio-Incubateur Eurasanté (2024)",
        "oui - Eurasanté (Bio-Incubateur)",
        "eurasante.com"
    ),
    (
        "Happlyz Medical",
        "contact@happlyz.com",
        "https://fr.linkedin.com/company/happlyzmedical",
        "Vanessa Lesobre (CEO), Maxime Berriot (co-fondateur)",
        "Santé / MedTech - Rééducation respiratoire par gamification (solution LORIO)",
        "En incubation - Bio-Incubateur Eurasanté (2024)",
        "oui - Eurasanté (Bio-Incubateur)",
        "eurasante.com / happlyzmedical.com"
    ),
    (
        "Myodev",
        "francois.ottavi@myodev.com",
        "https://www.linkedin.com/in/francois-ottavi-525b28157",
        "François Ottavi (CEO & Fondateur)",
        "Santé / e-santé - Assistant numérique de rééducation musculaire (vêtement connecté EMG)",
        "Accompagné par Eurasanté (2024)",
        "oui - Eurasanté (partenaire)",
        "eurasante.com / myodev.cloud"
    ),
    (
        "CELEOS",
        "",
        "https://www.eurasante.com/entreprises/celeos/",
        "Philippe Saudemont (CEO & Co-fondateur)",
        "Santé / MedTech - Dispositif médical SpiderMass + IA pour analyser les marges de résection en chirurgie oncologique",
        "En incubation - Bio-Incubateur Eurasanté (depuis 2023, prix Pépite HODEFI 2024)",
        "oui - Eurasanté (Bio-Incubateur)",
        "eurasante.com"
    ),
    (
        "Omniscience",
        "contact@omniscience.cloud",
        "https://www.linkedin.com/company/omnisciencecloud",
        "",
        "Santé / e-santé - SaaS de gestion des essais cliniques (collecte, monitoring, analyse de données patients)",
        "En incubation - Bio-Incubateur Eurasanté (2024)",
        "oui - Eurasanté (Bio-Incubateur)",
        "eurasante.com / omniscience.cloud"
    ),
    (
        "Sterion",
        "",
        "https://fr.linkedin.com/company/sterion",
        "",
        "Santé - Gestion, désinfection, stérilisation et recyclage des déchets médicaux",
        "Bio-Start Eurasanté (2024)",
        "oui - Eurasanté (Bio-Start 2024)",
        "eurasante.com"
    ),
    (
        "Keenamics",
        "",
        "",
        "Pascal Breton, Guillaume Valenzuela",
        "Santé / MedTech - Ergomètre portable pour rééducation de la cheville (recherche UPHF)",
        "En incubation - Bio-Incubateur Eurasanté + Vivalley (2024)",
        "oui - Eurasanté (Bio-Incubateur) + Vivalley",
        "eurasante.com / vivalley-campus.fr"
    ),

    # ======= EURALIMENTAIRE =======
    (
        "BellyCare",
        "",
        "https://fr.linkedin.com/company/euralimentaire",
        "Melissa Bertin, Maxime Blanc, Rémy Degrolard",
        "Alimentation - Repas frais certifiés Low FODMAP (sans gluten, sans lactose) pour troubles digestifs",
        "En incubation - Euralimentaire (actif 2024)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com"
    ),
    (
        "ENAKO",
        "",
        "https://enako.fr",
        "Antoine (fondateur)",
        "Alimentation / Foodtech - Boisson énergisante naturelle (thé vert, ginseng, fruits, sans additif)",
        "En incubation - Euralimentaire (2024-2025)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com / enako.fr"
    ),
    (
        "CHAMPITERNEL (Fun Guy Foods)",
        "",
        "",
        "",
        "Alimentation / Foodtech - Substituts de viande à base de champignons",
        "En incubation - Euralimentaire (2024)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com"
    ),
    (
        "DELIBUS (Saveur Voyage)",
        "",
        "",
        "",
        "Alimentation - Service de restauration pour autocars et livraison de paniers-repas",
        "En incubation - Euralimentaire (2024)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com"
    ),
    (
        "BREADSTHYME (La Sandwicherie Libanaise)",
        "",
        "",
        "",
        "Alimentation - Sandwicherie libanaise",
        "En incubation - Euralimentaire (2024)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com"
    ),
    (
        "MAISON SEDDI",
        "",
        "",
        "",
        "Alimentation - Confitures et tartinades (alternatives au sucre)",
        "En incubation - Euralimentaire (2024)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com"
    ),
    (
        "SEANOVATION",
        "",
        "",
        "",
        "Alimentation - Produits issus de la mer / innovation alimentaire",
        "En incubation - Euralimentaire (2024)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com"
    ),
    (
        "WE ARE APEX (Smart Fuel)",
        "",
        "",
        "",
        "Alimentation / Sport - Compléments nutritionnels pour sportifs",
        "En incubation - Euralimentaire (2024)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com"
    ),
    (
        "APPLOS (M.I.TEA)",
        "",
        "",
        "",
        "Alimentation / Boissons - Boissons sans sucre à base de thé de montagne grec",
        "En incubation - Euralimentaire (2024)",
        "oui - Eurasanté (Euralimentaire)",
        "euralimentaire.com"
    ),

    # ======= EURASENIOR =======
    (
        "COOZZIE",
        "",
        "",
        "Philippe DELSOL (fondateur)",
        "Silver économie - Services / innovation pour seniors (fondée décembre 2024)",
        "En incubation - Eurasenior (depuis déc. 2024)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "AUGIAS FRANCE SOLUTIONS",
        "",
        "",
        "",
        "Silver économie - Solutions innovantes pour le bien-vieillir",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "BOUGER",
        "",
        "",
        "",
        "Silver économie - Mobilité et activité physique pour seniors",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "CARENOVA - PAYELO",
        "",
        "",
        "",
        "Silver économie - Services de soins et paiement pour seniors",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "Hapigo (ex Cohevio)",
        "",
        "",
        "",
        "Silver économie - Habitat et lien social pour seniors",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "KOPPELIA",
        "",
        "",
        "",
        "Silver économie - Bien-être et lien social pour personnes âgées",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "L'Adaptelier",
        "",
        "",
        "",
        "Silver économie - Adaptation du logement pour seniors et personnes à mobilité réduite",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "PUPIL",
        "",
        "",
        "",
        "Silver économie - Technologie / services innovants pour seniors",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "REUMED (LIBEL'UP)",
        "",
        "",
        "",
        "Silver économie / Santé - Rééducation médicale pour seniors",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "UNVOID",
        "",
        "",
        "",
        "Silver économie - Innovation contre l'isolement des personnes âgées",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "Upeak",
        "",
        "",
        "",
        "Silver économie - Bien-vieillir et prévention santé pour seniors",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),
    (
        "WheelMove",
        "",
        "",
        "",
        "Silver économie / MedTech - Dispositif de mobilité innovant pour personnes en fauteuil roulant",
        "En incubation - Eurasenior (2024-2025)",
        "oui - Eurasanté (Eurasenior)",
        "eurasenior.fr"
    ),

    # ======= VIVALLEY =======
    (
        "PLAYERLYNK",
        "",
        "",
        "",
        "Sport / Digital - Plateforme numérique de recrutement entre clubs de basketball et agents de joueurs",
        "En incubation - Vivalley by Eurasanté (depuis juil. 2024)",
        "oui - Eurasanté (Vivalley)",
        "vivalley-campus.fr"
    ),
    (
        "HERA",
        "",
        "",
        "Louise Planchon (fondatrice)",
        "Sport / Santé / Femmes - Accompagnement holistique pré et post-partum (santé, sport, bien-être, nutrition)",
        "En incubation - Vivalley by Eurasanté (depuis juil. 2024)",
        "oui - Eurasanté (Vivalley)",
        "vivalley-campus.fr"
    ),
    (
        "KOBI",
        "",
        "",
        "Sofiane Laurent (CEO)",
        "Sport / Santé - Innovation sport & bien-être",
        "En incubation - Vivalley by Eurasanté (2024)",
        "oui - Eurasanté (Vivalley)",
        "vivalley-campus.fr"
    ),
    (
        "Setting Events",
        "",
        "",
        "",
        "Sport / Marketing - Valorisation de marque et gestion des partenariats sportifs",
        "En incubation - Vivalley by Eurasanté (automne 2025)",
        "oui - Eurasanté (Vivalley)",
        "vivalley-campus.fr"
    ),
    (
        "Performeo",
        "",
        "",
        "",
        "Sport / Formation - Formation et expertise dans l'industrie du fitness",
        "En incubation - Vivalley by Eurasanté (2025)",
        "oui - Eurasanté (Vivalley)",
        "vivalley-campus.fr"
    ),
    (
        "Vipali",
        "",
        "",
        "",
        "Santé / Prévention - Prévention santé personnalisée",
        "En incubation - Vivalley by Eurasanté (2025)",
        "oui - Eurasanté (Vivalley)",
        "vivalley-campus.fr"
    ),
]

start_row = 282
for i, (nom, email, linkedin, founder, offre, stade, accompagnee, trouvee_sur) in enumerate(startups):
    row = start_row + i
    ws.cell(row=row, column=2).value = nom
    ws.cell(row=row, column=3).value = email if email else None
    ws.cell(row=row, column=4).value = linkedin if linkedin else None
    ws.cell(row=row, column=5).value = founder if founder else None
    ws.cell(row=row, column=6).value = offre
    ws.cell(row=row, column=7).value = stade
    ws.cell(row=row, column=10).value = accompagnee
    ws.cell(row=row, column=11).value = trouvee_sur

wb.save('c:/Users/gaiam/Desktop/Liquid-/Liquid+ Project/Sélection des startups/processus de sélection des startups GAIA v2_TEMP.xlsx')
print(f"Done! Added {len(startups)} startups starting from row {start_row}.")
print(f"Last row used: {start_row + len(startups) - 1}")
